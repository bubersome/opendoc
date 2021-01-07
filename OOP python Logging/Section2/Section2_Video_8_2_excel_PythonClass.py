'''
Created on Jun 19, 2019
@author: Burkhard A. Meier
'''








from openpyxl import Workbook

def write_to_excel(class_data):
    wb = Workbook()          
    ws = wb.active    
    ws.title = 'Data'       
    row = col = 1
  
    for key, val in class_data.items():
        if isinstance(val, dict):
            ws.cell(column=col, row=row, value=key)                         # write key
            for key1, val1 in val.items():
                if isinstance(val1, dict):
                    ws.cell(column=col+1, row=row, value=key1)              # write key
                    col += 1
                    for key2, val2 in val1.items():
                        ws.cell(column=col+1, row=row, value=key2)          # write key
                        ws.cell(column=col+2, row=row, value=val2)          # write value into next col
                        row += 1                       
                else:
                    ws.cell(column=col+1, row=row, value=key1)          # write key
                    ws.cell(column=col+2, row=row, value=val1)          # write value into next col
                    row += 1
        else:
            ws.cell(column=col, row=row, value=key)               # write key
            ws.cell(column=col+1, row=row, value=val)             # write value into next col
            row += 1
    
    wb.save('PythonClass.xlsx')         # save workbook with a file name
    
    

class PythonClass(object):
    def __init__(self, name='default name', items=None):      
        self.name = name    
        self.items = items                        

    def get_name(self):
        return f"Name is: {self.name}"                

    def get_items(self):
        return f"Items are: {self.items}"
            
    def get_name_items_dict(self):
        return dict(name = self.name,
                    items = self.items)
        
        
some_people = {'people': {'First': 'Bob', 'Second': 'Bill', 'Third': 'unknown'}}
data_dict_class = PythonClass('data_dict', some_people)
class_data = data_dict_class.get_name_items_dict()

print(class_data)
write_to_excel(class_data)






