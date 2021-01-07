'''
Created on Jun 19, 2019
@author: Burkhard A. Meier
'''







'''
see: https://openpyxl.readthedocs.io/en/stable/tutorial.html
'''
from openpyxl import Workbook, load_workbook
#-----------------------------------------------------------------
wb = Workbook()                     # create new Workbook
print(wb)

ws = wb.active                      # get the active worksheet
print(ws)
ws.title = 'Data'                   # give worksheet a title

ws['A1'] = 'Test'                   # write data to cell A1

print(ws['A1'].value)               # get the cell value

wb.save('PythonClass.xlsx')         # save workbook under a name

#-----------------------------------------------------------------

wb1 = load_workbook('PythonClass.xlsx')     # load existing workbook file

ws1 = wb1['Data']                           # get worksheet by name

print(ws1['A1'].value)                      # print cell value
















